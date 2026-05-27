"""Data lane: materialize tabular sources into Parquet under ``data/``.

Used when a document is tagged ``data`` (see :mod:`services.paths`). Supports
the eight data-class extensions:

* ``.csv`` / ``.tsv``       -> one Parquet per file
* ``.xlsx`` / ``.xls``      -> one Parquet per sheet
* ``.json``                 -> Parquet if top-level is array-of-objects
* ``.yaml`` / ``.yml``      -> Parquet if top-level is array-of-objects
* ``.xml``                  -> Parquet if root has uniform child elements

Best-effort: when a file isn't naturally tabular (e.g. nested JSON config),
no Parquet is produced and a structured reason is returned. The metadata-only
Markdown preview in :mod:`services.text_pipeline` still gets written either way.
"""

from __future__ import annotations

import csv as _csv
import json as _json
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any


class DataPipelineError(RuntimeError):
    pass


@dataclass
class DataTableArtifact:
    table_name: str               # stable handle (also used in MCP tools)
    sheet: str | None             # xlsx/xls sheet name; None otherwise
    parquet_path: Path            # absolute path on disk
    columns: list[str]
    row_count: int


_SLUG_RE = re.compile(r"[^A-Za-z0-9_]+")


def _slug(s: str) -> str:
    s = _SLUG_RE.sub("_", s).strip("_")
    return s or "table"


def _table_name(doc_id: str, sheet: str | None) -> str:
    base = _slug(doc_id)
    return base if sheet is None else f"{base}__{_slug(sheet)}"


def _parquet_target(data_dir: Path, source_rel: str, sheet: str | None) -> Path:
    """Mirror the documents/ tree under data/, replacing the suffix.

    Multi-sheet files get one Parquet per sheet: ``foo.xlsx`` ->
    ``foo.<sheet>.parquet``.
    """
    src = Path(source_rel)
    if sheet is None:
        rel = src.with_suffix(".parquet")
    else:
        rel = src.with_suffix("") .with_name(f"{src.stem}.{_slug(sheet)}.parquet")
    return data_dir / rel


def _write_parquet(out_path: Path, columns: list[str], rows: list[list[Any]]) -> int:
    """Write rows to Parquet using pyarrow. All cells are coerced to string to
    keep the writer schema-stable across heterogeneous tabular sources.
    """
    try:
        import pyarrow as pa  # type: ignore
        import pyarrow.parquet as pq  # type: ignore
    except ImportError as e:
        raise DataPipelineError(
            "pyarrow not installed: reinstall the project with `pip install -e .`"
        ) from e

    n_cols = len(columns)
    str_columns: list[list[str | None]] = [[] for _ in range(n_cols)]
    for r in rows:
        for i in range(n_cols):
            v = r[i] if i < len(r) else None
            if v is None:
                str_columns[i].append(None)
            else:
                str_columns[i].append(str(v))

    table = pa.table({columns[i]: str_columns[i] for i in range(n_cols)})
    out_path.parent.mkdir(parents=True, exist_ok=True)
    pq.write_table(table, out_path, compression="zstd")
    return len(rows)


# ---------------------------------------------------------------------------
# Per-format readers (return list of (sheet|None, columns, rows))
# ---------------------------------------------------------------------------


def _read_csv(path: Path) -> list[tuple[str | None, list[str], list[list[Any]]]]:
    delim = "\t" if path.suffix.lower() == ".tsv" else ","
    with path.open("r", encoding="utf-8-sig", newline="", errors="replace") as f:
        reader = _csv.reader(f, delimiter=delim)
        header = next(reader, None)
        if not header:
            raise DataPipelineError("csv/tsv has no header row")
        columns = [h or f"col{i}" for i, h in enumerate(header)]
        rows = [list(r) for r in reader]
    return [(None, columns, rows)]


def _read_xlsx(path: Path) -> list[tuple[str | None, list[str], list[list[Any]]]]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        raise DataPipelineError("openpyxl not installed") from e
    out: list[tuple[str | None, list[str], list[list[Any]]]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for name in wb.sheetnames:
            ws = wb[name]
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row is None:
                continue
            columns = [
                str(c) if c is not None and str(c) else f"col{i}"
                for i, c in enumerate(header_row)
            ]
            data: list[list[Any]] = []
            for r in rows_iter:
                data.append(list(r))
            # Skip completely empty sheets (header-only is fine).
            out.append((name, columns, data))
    finally:
        wb.close()
    if not out:
        raise DataPipelineError("xlsx has no non-empty sheets")
    return out


def _read_xls(path: Path) -> list[tuple[str | None, list[str], list[list[Any]]]]:
    try:
        import xlrd  # type: ignore
    except ImportError as e:
        raise DataPipelineError("xlrd not installed") from e
    out: list[tuple[str | None, list[str], list[list[Any]]]] = []
    book = xlrd.open_workbook(str(path), on_demand=True)
    for name in book.sheet_names():
        sh = book.sheet_by_name(name)
        if sh.nrows == 0:
            continue
        header = sh.row_values(0)
        columns = [
            str(c) if c is not None and str(c) else f"col{i}"
            for i, c in enumerate(header)
        ]
        data = [sh.row_values(i) for i in range(1, sh.nrows)]
        out.append((name, columns, data))
    if not out:
        raise DataPipelineError("xls has no non-empty sheets")
    return out


def _read_json(path: Path) -> list[tuple[str | None, list[str], list[list[Any]]]]:
    raw = path.read_text(encoding="utf-8", errors="replace")
    try:
        data = _json.loads(raw)
    except Exception as e:  # noqa: BLE001
        raise DataPipelineError(f"json parse failed: {e}") from e
    if not isinstance(data, list) or not data or not all(isinstance(x, dict) for x in data):
        raise DataPipelineError("json is not an array-of-objects; cannot tabularize")
    columns: list[str] = []
    seen: set[str] = set()
    for item in data:
        for k in item.keys():
            if k not in seen:
                seen.add(k)
                columns.append(k)
    rows: list[list[Any]] = []
    for item in data:
        rows.append([item.get(c) for c in columns])
    return [(None, columns, rows)]


def _read_yaml(path: Path) -> list[tuple[str | None, list[str], list[list[Any]]]]:
    try:
        import yaml as _yaml  # type: ignore
    except ImportError as e:
        raise DataPipelineError("pyyaml not installed") from e
    try:
        data = _yaml.safe_load(path.read_text(encoding="utf-8", errors="replace"))
    except Exception as e:  # noqa: BLE001
        raise DataPipelineError(f"yaml parse failed: {e}") from e
    if not isinstance(data, list) or not data or not all(isinstance(x, dict) for x in data):
        raise DataPipelineError("yaml is not an array-of-objects; cannot tabularize")
    columns: list[str] = []
    seen: set[str] = set()
    for item in data:
        for k in item.keys():
            if k not in seen:
                seen.add(k)
                columns.append(k)
    rows = [[item.get(c) for c in columns] for item in data]
    return [(None, columns, rows)]


def _read_xml(path: Path) -> list[tuple[str | None, list[str], list[list[Any]]]]:
    from xml.etree import ElementTree as ET

    try:
        root = ET.parse(str(path)).getroot()
    except Exception as e:  # noqa: BLE001
        raise DataPipelineError(f"xml parse failed: {e}") from e

    children = list(root)
    if not children:
        raise DataPipelineError("xml root has no child elements")

    # Require uniform child tag for a clean tabular shape.
    first_tag = children[0].tag
    uniform = all(c.tag == first_tag for c in children)
    if not uniform:
        raise DataPipelineError("xml children have mixed tags; cannot tabularize")

    columns: list[str] = []
    seen: set[str] = set()
    for el in children:
        for k in el.attrib.keys():
            if k not in seen:
                seen.add(k)
                columns.append(k)
        for sub in list(el):
            if sub.tag not in seen:
                seen.add(sub.tag)
                columns.append(sub.tag)
    # Always include the element text as ``_text`` if any child element uses it.
    has_text = any((el.text or "").strip() for el in children)
    if has_text and "_text" not in seen:
        columns.append("_text")
        seen.add("_text")

    rows: list[list[Any]] = []
    for el in children:
        row: list[Any] = []
        for c in columns:
            if c == "_text":
                row.append((el.text or "").strip() or None)
            elif c in el.attrib:
                row.append(el.attrib.get(c))
            else:
                sub = el.find(c)
                row.append((sub.text or "").strip() if sub is not None and sub.text else None)
        rows.append(row)
    return [(None, columns, rows)]


_READERS: dict[str, Any] = {
    ".csv": _read_csv,
    ".tsv": _read_csv,
    ".xlsx": _read_xlsx,
    ".xls": _read_xls,
    ".json": _read_json,
    ".yaml": _read_yaml,
    ".yml": _read_yaml,
    ".xml": _read_xml,
}


# ---------------------------------------------------------------------------
# Public entry point
# ---------------------------------------------------------------------------


def build_parquet_artifacts(
    src_path: Path,
    *,
    doc_id: str,
    source_rel: str,
    data_dir: Path,
) -> list[DataTableArtifact]:
    """Convert a single data-class source file into one or more Parquet files.

    Returns the list of produced artifacts; raises :class:`DataPipelineError`
    when the file cannot be tabularized.
    """
    suffix = src_path.suffix.lower()
    reader = _READERS.get(suffix)
    if reader is None:
        raise DataPipelineError(f"no data pipeline reader for {suffix}")

    parts = reader(src_path)

    artifacts: list[DataTableArtifact] = []
    for sheet, columns, rows in parts:
        out = _parquet_target(data_dir, source_rel, sheet)
        n = _write_parquet(out, columns, rows)
        artifacts.append(
            DataTableArtifact(
                table_name=_table_name(doc_id, sheet),
                sheet=sheet,
                parquet_path=out,
                columns=list(columns),
                row_count=n,
            )
        )
    return artifacts
