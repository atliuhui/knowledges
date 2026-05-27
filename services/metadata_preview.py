"""Metadata-only Markdown previews.

Used when a source file is NOT eligible for full content extraction (e.g.
data-class files without the ``text`` tag). The output starts with a YAML
front matter block carrying the document's metadata fields, followed by a
short schema preview (columns + sample rows) for tabular formats. The body
deliberately does not include full content, so it stays small while remaining
useful for full-text ("does this file even exist?") and Agent display
purposes.

See ``services/text_pipeline.py`` for the dispatcher entry point.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from ._text_base import ConversionResult, _normalize_text


def _format_yaml_scalar(v: Any) -> str:
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "true" if v else "false"
    if isinstance(v, (int, float)):
        return str(v)
    s = str(v)
    # Always quote to keep YAML happy and preserve original text.
    return '"' + s.replace("\\", "\\\\").replace('"', '\\"') + '"'


def _schema_preview_csv(path: Path, sample_rows: int = 5) -> list[str]:
    import csv as _csv
    out: list[str] = []
    delim = "\t" if path.suffix.lower() == ".tsv" else ","
    try:
        with path.open("r", encoding="utf-8-sig", newline="", errors="replace") as f:
            reader = _csv.reader(f, delimiter=delim)
            header = next(reader, None) or []
            out.append("## Columns")
            out.append("")
            if header:
                out.extend(f"- `{h}`" for h in header)
            else:
                out.append("(no header)")
            out.append("")
            samples: list[list[str]] = []
            for _ in range(sample_rows):
                row = next(reader, None)
                if row is None:
                    break
                samples.append(row)
            if samples and header:
                out.append("## Sample (first {} rows)".format(len(samples)))
                out.append("")
                out.append("| " + " | ".join(header) + " |")
                out.append("|" + "|".join(["---"] * len(header)) + "|")
                for r in samples:
                    cells = list(r) + [""] * max(0, len(header) - len(r))
                    cells = [c.replace("|", "\\|").replace("\n", " ") for c in cells[:len(header)]]
                    out.append("| " + " | ".join(cells) + " |")
                out.append("")
    except Exception as e:  # noqa: BLE001
        out.append(f"<!-- schema preview failed: {e} -->")
    return out


def _schema_preview_xlsx(path: Path, sample_rows: int = 5) -> list[str]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        return ["<!-- openpyxl not available; cannot preview xlsx schema -->"]
    out: list[str] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as e:  # noqa: BLE001
        return [f"<!-- xlsx open failed: {e} -->"]
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            out.append(f"## Sheet `{sheet_name}`")
            out.append("")
            rows_iter = ws.iter_rows(values_only=True)
            header_row = next(rows_iter, None)
            if header_row is None:
                out.append("(empty sheet)")
                out.append("")
                continue
            header = ["" if c is None else str(c) for c in header_row]
            out.append("### Columns")
            out.append("")
            out.extend(f"- `{h}`" if h else "- *(blank)*" for h in header)
            out.append("")
            samples: list[list[str]] = []
            for _ in range(sample_rows):
                r = next(rows_iter, None)
                if r is None:
                    break
                samples.append(["" if c is None else str(c) for c in r])
            if samples:
                out.append(f"### Sample (first {len(samples)} rows)")
                out.append("")
                out.append("| " + " | ".join(header) + " |")
                out.append("|" + "|".join(["---"] * len(header)) + "|")
                for s in samples:
                    cells = s + [""] * max(0, len(header) - len(s))
                    cells = [c.replace("|", "\\|").replace("\n", " ") for c in cells[:len(header)]]
                    out.append("| " + " | ".join(cells) + " |")
                out.append("")
    finally:
        wb.close()
    return out


def _schema_preview_xls(path: Path, sample_rows: int = 5) -> list[str]:
    try:
        import xlrd  # type: ignore
    except ImportError:
        return ["<!-- xlrd not available; cannot preview xls schema -->"]
    out: list[str] = []
    try:
        book = xlrd.open_workbook(str(path), on_demand=True)
    except Exception as e:  # noqa: BLE001
        return [f"<!-- xls open failed: {e} -->"]
    for sheet_name in book.sheet_names():
        sh = book.sheet_by_name(sheet_name)
        out.append(f"## Sheet `{sheet_name}`")
        out.append("")
        if sh.nrows == 0:
            out.append("(empty sheet)")
            out.append("")
            continue
        header = ["" if v is None else str(v) for v in sh.row_values(0)]
        out.append("### Columns")
        out.append("")
        out.extend(f"- `{h}`" if h else "- *(blank)*" for h in header)
        out.append("")
        n = min(sample_rows, sh.nrows - 1)
        if n > 0:
            out.append(f"### Sample (first {n} rows)")
            out.append("")
            out.append("| " + " | ".join(header) + " |")
            out.append("|" + "|".join(["---"] * len(header)) + "|")
            for i in range(1, 1 + n):
                vals = ["" if v is None else str(v) for v in sh.row_values(i)]
                cells = vals + [""] * max(0, len(header) - len(vals))
                cells = [c.replace("|", "\\|").replace("\n", " ") for c in cells[:len(header)]]
                out.append("| " + " | ".join(cells) + " |")
            out.append("")
    return out


def _schema_preview_json(path: Path, sample_rows: int = 5) -> list[str]:
    import json as _json
    out: list[str] = []
    try:
        data = _json.loads(path.read_text(encoding="utf-8", errors="replace"))
    except Exception as e:  # noqa: BLE001
        return [f"<!-- json parse failed: {e} -->"]
    if isinstance(data, list) and data and isinstance(data[0], dict):
        keys: list[str] = []
        seen: set[str] = set()
        for item in data[: max(50, sample_rows)]:
            if isinstance(item, dict):
                for k in item.keys():
                    if k not in seen:
                        seen.add(k)
                        keys.append(k)
        out.append(f"## Array of {len(data)} records")
        out.append("")
        out.append("### Columns (union of first records)")
        out.append("")
        out.extend(f"- `{k}`" for k in keys)
        out.append("")
        samples = data[:sample_rows]
        if samples and keys:
            out.append(f"### Sample (first {len(samples)} records)")
            out.append("")
            out.append("| " + " | ".join(keys) + " |")
            out.append("|" + "|".join(["---"] * len(keys)) + "|")
            for item in samples:
                if not isinstance(item, dict):
                    continue
                cells = []
                for k in keys:
                    v = item.get(k, "")
                    cell = "" if v is None else str(v)
                    cell = cell.replace("|", "\\|").replace("\n", " ")
                    if len(cell) > 80:
                        cell = cell[:77] + "..."
                    cells.append(cell)
                out.append("| " + " | ".join(cells) + " |")
            out.append("")
    else:
        out.append("## Structure")
        out.append("")
        out.append(f"- Top-level type: `{type(data).__name__}`")
        if isinstance(data, dict):
            out.append(f"- Top-level keys: {', '.join(f'`{k}`' for k in list(data.keys())[:30])}")
        elif isinstance(data, list):
            out.append(f"- Length: {len(data)}")
        out.append("")
    return out


def _schema_preview_xml(path: Path, sample_rows: int = 5) -> list[str]:
    from xml.etree import ElementTree as ET
    try:
        root = ET.parse(str(path)).getroot()
    except Exception as e:  # noqa: BLE001
        return [f"<!-- xml parse failed: {e} -->"]
    out: list[str] = ["## Structure", ""]
    out.append(f"- Root tag: `{root.tag}`")
    children_tags: list[str] = []
    seen: set[str] = set()
    for child in list(root)[:200]:
        if child.tag not in seen:
            seen.add(child.tag)
            children_tags.append(child.tag)
    if children_tags:
        out.append(f"- Child tags: {', '.join(f'`{t}`' for t in children_tags[:30])}")
    out.append("")
    samples = list(root)[:sample_rows]
    if samples:
        out.append(f"### Sample (first {len(samples)} child elements)")
        out.append("")
        for el in samples:
            attrs = " ".join(f'{k}="{v}"' for k, v in el.attrib.items())
            text = (el.text or "").strip().replace("\n", " ")
            if len(text) > 80:
                text = text[:77] + "..."
            head = f"<{el.tag}" + (f" {attrs}" if attrs else "") + ">"
            out.append(f"- {head} {text}".rstrip())
        out.append("")
    return out


def _schema_preview_yaml(path: Path, sample_rows: int = 5) -> list[str]:
    try:
        import yaml as _yaml  # type: ignore
    except ImportError:
        return ["<!-- pyyaml not available; cannot preview yaml schema -->"]
    try:
        data = _yaml.safe_load(path.read_text(encoding="utf-8", errors="replace"))
    except Exception as e:  # noqa: BLE001
        return [f"<!-- yaml parse failed: {e} -->"]
    out: list[str] = ["## Structure", ""]
    out.append(f"- Top-level type: `{type(data).__name__}`")
    if isinstance(data, dict):
        out.append(f"- Top-level keys: {', '.join(f'`{k}`' for k in list(data.keys())[:30])}")
    elif isinstance(data, list):
        out.append(f"- Length: {len(data)}")
        if data and isinstance(data[0], dict):
            keys: list[str] = []
            seen: set[str] = set()
            for item in data[:50]:
                if isinstance(item, dict):
                    for k in item.keys():
                        if k not in seen:
                            seen.add(k)
                            keys.append(k)
            if keys:
                out.append(f"- First-item keys: {', '.join(f'`{k}`' for k in keys[:30])}")
    out.append("")
    return out


_SCHEMA_PREVIEWERS: dict[str, Any] = {
    ".csv": _schema_preview_csv,
    ".tsv": _schema_preview_csv,
    ".xlsx": _schema_preview_xlsx,
    ".xls": _schema_preview_xls,
    ".json": _schema_preview_json,
    ".xml": _schema_preview_xml,
    ".yaml": _schema_preview_yaml,
    ".yml": _schema_preview_yaml,
}


def convert_metadata_only(
    path: Path,
    *,
    source_rel: str,
    title: str = "",
    doc_type: str = "",
    tags: str = "",
    confidentiality: str = "",
    status: str = "",
    notes: str = "",
    source_size: str | int = "",
    source_mtime: str = "",
    sample_rows: int = 5,
) -> ConversionResult:
    """Build a *metadata-only* Markdown preview for a source file.

    Output layout::

        ---
        title: "..."
        source_path: "..."
        ...
        ---

        # <stem>

        ## Columns / ## Structure / ## Sheet `...`  (only for data-class files)
        ...
    """
    suffix = path.suffix.lower()
    fm: list[tuple[str, Any]] = [
        ("title", title or path.stem),
        ("source_path", source_rel),
        ("type", doc_type),
        ("tags", tags),
        ("confidentiality", confidentiality),
        ("status", status),
        ("size", source_size),
        ("mtime", source_mtime),
        ("suffix", suffix),
        ("notes", notes),
    ]
    lines: list[str] = ["---"]
    for k, v in fm:
        lines.append(f"{k}: {_format_yaml_scalar(v)}")
    lines.append("---")
    lines.append("")
    lines.append(f"# {title or path.stem}")
    lines.append("")
    lines.append(
        "> Metadata-only preview. Full content is not extracted because this "
        "document is not tagged for text indexing."
    )
    lines.append("")

    previewer = _SCHEMA_PREVIEWERS.get(suffix)
    if previewer is not None:
        try:
            preview = previewer(path, sample_rows)
        except Exception as e:  # noqa: BLE001
            preview = [f"<!-- schema preview failed: {e} -->"]
        lines.extend(preview)

    return ConversionResult(_normalize_text("\n".join(lines)), ".md")
